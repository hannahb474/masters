"""
study2_time_and_bvalue_reduction.py
Study 2 (report section 6): Reducing b-values and diffusion times in patients
and controls.

Every shortened acquisition (a subset of the 5 diffusion times x a subset of the
4 b-value shells = 465 protocols) is evaluated. Tensors are fit PER diffusion
time (that block's b0 + retained shells, all averages); protocols keeping several
diffusion times pool them by averaging. For each protocol it reports:

  * group-mean MD and FA (controls vs DMD),
  * a group comparison (Mann-Whitney U + Cohen's d + rank-biserial r),
  * the Cramer-Rao lower bound (CRLB) on MD and FA precision,
  * the voxelwise RMSE of MD and FA against the FULL acquisition (reference).

Caveat (5 controls vs 4 patients): the smallest two-sided Mann-Whitney p is
0.0159, so p < 0.05 is near-binary. Read effect sizes and the "matches full
protocol" verdict, not p alone.

Outputs: shortscan_results.xlsx, protocol_results.csv, summary.txt, heatmaps.

Requires the local modules dti_scheme.py (acquisition scheme) and crlb.py.

    python study2_time_and_bvalue_reduction.py --data-root /path/to/study \
        --bval dwi.bval --bvec dwi.bvec --outdir results/
    python study2_time_and_bvalue_reduction.py --selftest --outdir selftest_out/
"""

from __future__ import annotations
import argparse
import glob
import itertools
import os
import numpy as np
import pandas as pd

from dti_scheme import AcquisitionScheme, load_scheme
import crlb as crlb_mod

# Subject folder names under the data root (edit to match your dataset).
DEFAULT_HEALTHY_FOLDERS = ["control_01", "control_02", "control_03", "control_04", "control_05"]
DEFAULT_PATIENT_FOLDERS = ["patient_01", "patient_02", "patient_03", "patient_04"]


def _import_heavy():
    import nibabel as nib
    from dipy.core.gradients import gradient_table
    from dipy.reconst.dti import TensorModel
    from scipy import stats
    return nib, gradient_table, TensorModel, stats


# ---------------------------------------------------------------------- #
# Build the subject manifest (subject_id, group, dwi, mask) from folders
# ---------------------------------------------------------------------- #
def build_manifest_from_root(data_root, healthy_folders, patient_folders,
                             healthy_label, patient_label,
                             dwi_glob="*dwi*.nii*", mask_glob="*mask*.nii*"):
    rows = []
    for folders, label in ((healthy_folders, healthy_label),
                           (patient_folders, patient_label)):
        for f in folders:
            fdir = os.path.join(data_root, f)
            dwis = sorted(glob.glob(os.path.join(fdir, dwi_glob)))
            masks = sorted(glob.glob(os.path.join(fdir, mask_glob)))
            if not dwis:
                raise FileNotFoundError(f"No DWI matching '{dwi_glob}' in {fdir}")
            if not masks:
                raise FileNotFoundError(f"No mask matching '{mask_glob}' in {fdir}")
            rows.append(dict(subject_id=f, group=label, dwi=dwis[0], mask=masks[0]))
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------- #
# Noise SD from repeated averages of the same (block, shell, direction)
# ---------------------------------------------------------------------- #
def estimate_sigma_from_repeats(dwi, mask_bool, scheme):
    bvecs = scheme.bvecs
    group_vars = []
    for blk in scheme.blocks:
        buckets = {}
        for b, idxs in blk.shells.items():
            for i in idxs:
                key = (b, tuple(np.round(bvecs[:, i], 3)))
                buckets.setdefault(key, []).append(i)
        for idxs in buckets.values():
            if len(idxs) < 2:
                continue
            v = np.var(dwi[..., idxs][mask_bool], axis=1, ddof=1)   # per-voxel variance across repeats
            v = v[np.isfinite(v)]
            if v.size:
                group_vars.append(np.mean(v))
    return float(np.sqrt(np.median(group_vars))) if group_vars else np.nan


# ---------------------------------------------------------------------- #
# Ground-truth tensor per subject/diffusion time (weighted log-linear fit)
# ---------------------------------------------------------------------- #
def loglin_tensor_fit(bvals, bvecs, signal):
    """Weighted log-linear DTI fit on one averaged signal vector.
    Returns (params[Dxx,Dyy,Dzz,Dxy,Dxz,Dyz], S0)."""
    bvals = np.asarray(bvals, float).ravel()
    bvecs = np.asarray(bvecs, float)
    if bvecs.shape[0] != 3:
        bvecs = bvecs.T
    gx, gy, gz = bvecs
    X = np.column_stack([
        np.ones_like(bvals),
        -bvals * gx * gx, -bvals * gy * gy, -bvals * gz * gz,
        -2 * bvals * gx * gy, -2 * bvals * gx * gz, -2 * bvals * gy * gz,
    ])
    s = np.clip(np.asarray(signal, float).ravel(), 1e-6, None)
    y = np.log(s)
    w = s ** 2                                   # WLS weights ~ S^2
    W = np.diag(w)
    try:
        beta = np.linalg.solve(X.T @ W @ X, X.T @ W @ y)
    except np.linalg.LinAlgError:
        beta, *_ = np.linalg.lstsq(np.sqrt(w)[:, None] * X, np.sqrt(w) * y, rcond=None)
    return beta[1:7], float(np.exp(beta[0]))


# ---------------------------------------------------------------------- #
# Fit every (diffusion time, b-subset) for one subject
# ---------------------------------------------------------------------- #
def process_subject(dwi, mask, scheme, gradient_table, TensorModel,
                    fit_method="WLS", b0_threshold_fit=90.0):
    mask_bool = mask > 0
    bvalues = scheme.nominal_bvalues
    full_bsub = tuple(sorted(bvalues))
    sigma = estimate_sigma_from_repeats(dwi, mask_bool, scheme)

    b_subsets = []
    for r in range(1, len(bvalues) + 1):
        b_subsets.extend(itertools.combinations(bvalues, r))

    voxel, gt = {}, {}
    for blk in scheme.blocks:
        d = blk.delta
        # ground-truth tensor from all b-values on the ROI-mean signal (for CRLB)
        idx_full = scheme.select_indices([d], list(full_bsub), include_b0=True)[d]["indices"]
        bvals_f, bvecs_f = scheme.subset_gtab_arrays(idx_full)
        gt[d] = loglin_tensor_fit(bvals_f, bvecs_f, dwi[..., idx_full][mask_bool].mean(axis=0))

        # voxelwise MD/FA for each b-subset at this diffusion time
        for bsub in b_subsets:
            idx = scheme.select_indices([d], list(bsub), include_b0=True)[d]["indices"]
            bvals_s, bvecs_s = scheme.subset_gtab_arrays(idx)
            gtab = gradient_table(bvals_s, bvecs=bvecs_s, b0_threshold=b0_threshold_fit)
            fit = TensorModel(gtab, fit_method=fit_method).fit(
                dwi[..., idx].astype(np.float32), mask=mask_bool)
            md_vox = np.asarray(fit.md[mask_bool], float)
            fa_vox = np.clip(np.asarray(fit.fa[mask_bool], float), 0, 1)
            md_vox[~np.isfinite(md_vox)] = np.nan
            fa_vox[~np.isfinite(fa_vox)] = np.nan
            voxel[(d, tuple(sorted(bsub)))] = (md_vox.astype(np.float32), fa_vox.astype(np.float32))
    return dict(voxel=voxel, gt=gt, sigma=sigma, n_vox=int(mask_bool.sum()))


# ---------------------------------------------------------------------- #
# Group statistics + effect sizes
# ---------------------------------------------------------------------- #
def cohens_d(a, b):
    a, b = np.asarray(a, float), np.asarray(b, float)
    na, nb = len(a), len(b)
    if na < 2 or nb < 2:
        return np.nan
    sp2 = ((na - 1) * a.var(ddof=1) + (nb - 1) * b.var(ddof=1)) / (na + nb - 2)
    return 0.0 if sp2 == 0 else (a.mean() - b.mean()) / np.sqrt(sp2)


def compare_groups(vh, vp, stats):
    vh = np.asarray(vh, float); vp = np.asarray(vp, float)
    vh = vh[np.isfinite(vh)]; vp = vp[np.isfinite(vp)]
    out = dict(mean_h=np.nan, mean_p=np.nan, U=np.nan, p=np.nan,
               rank_biserial=np.nan, cohens_d=np.nan)
    if len(vh) < 2 or len(vp) < 2:
        if len(vh): out["mean_h"] = float(vh.mean())
        if len(vp): out["mean_p"] = float(vp.mean())
        return out
    out["mean_h"], out["mean_p"] = float(vh.mean()), float(vp.mean())
    try:
        U, p = stats.mannwhitneyu(vh, vp, alternative="two-sided")
        out["U"], out["p"] = float(U), float(p)
        out["rank_biserial"] = float(1 - 2 * U / (len(vh) * len(vp)))
    except ValueError:
        pass
    out["cohens_d"] = cohens_d(vh, vp)
    return out


# ---------------------------------------------------------------------- #
# Sweep every protocol and score it
# ---------------------------------------------------------------------- #
def run_analysis(manifest, scheme, healthy_label, patient_label, outdir,
                 fit_method="WLS", verbose=True):
    nib, gradient_table, TensorModel, stats = _import_heavy()
    os.makedirs(outdir, exist_ok=True)

    # fit every subject
    subj, group = {}, {}
    for _, row in manifest.iterrows():
        sid, grp = str(row["subject_id"]), str(row["group"])
        if verbose:
            print(f"[fit] {sid} ({grp}) ...", flush=True)
        dwi = np.asarray(nib.load(row["dwi"]).dataobj)
        mask = np.asarray(nib.load(row["mask"]).dataobj)
        if dwi.shape[-1] != scheme.N:
            raise ValueError(f"{sid}: {dwi.shape[-1]} volumes, scheme expects {scheme.N}")
        subj[sid] = process_subject(dwi, mask, scheme, gradient_table, TensorModel, fit_method=fit_method)
        group[sid] = grp
        if verbose:
            print(f"       {subj[sid]['n_vox']} ROI voxels, sigma={subj[sid]['sigma']:.3g}", flush=True)

    h_ids = [s for s in subj if group[s] == healthy_label]
    p_ids = [s for s in subj if group[s] == patient_label]
    full_bsub = tuple(sorted(scheme.nominal_bvalues))
    all_deltas = tuple(scheme.diffusion_times)

    def pooled_voxel(sid, delta_subset, bsub, which):
        v = subj[sid]["voxel"]; key = tuple(sorted(bsub))
        arrs = [v[(d, key)][0 if which == "MD" else 1] for d in delta_subset]
        return np.nanmean(np.vstack(arrs), axis=0)

    # each subject's FULL-acquisition voxelwise MD/FA (the reference for RMSE)
    full_vox = {sid: {"MD": pooled_voxel(sid, all_deltas, full_bsub, "MD"),
                      "FA": pooled_voxel(sid, all_deltas, full_bsub, "FA")} for sid in subj}

    def subject_roimean(sid, delta_subset, bsub, which):
        return float(np.nanmean(pooled_voxel(sid, delta_subset, bsub, which)))

    def subject_rmse_vs_full(sid, delta_subset, bsub, which):
        d = pooled_voxel(sid, delta_subset, bsub, which) - full_vox[sid][which]
        d = d[np.isfinite(d)]
        return float(np.sqrt(np.mean(d ** 2))) if d.size else np.nan

    def subject_crlb_var(sid, delta_subset, bsub, which):
        """Pooled CRLB variance across retained diffusion times (variance of the mean)."""
        sigma = subj[sid]["sigma"]
        if not np.isfinite(sigma) or sigma <= 0:
            return np.nan
        vars_ = []
        for d in delta_subset:
            params, S0 = subj[sid]["gt"][d]
            idx = scheme.select_indices([d], list(bsub), include_b0=True)[d]["indices"]
            bvals_s, bvecs_s = scheme.subset_gtab_arrays(idx)
            vmd, vfa = crlb_mod.crlb_md_fa(bvals_s, bvecs_s, params, S0, sigma)
            vars_.append(vmd if which == "MD" else vfa)
        vars_ = [v for v in vars_ if np.isfinite(v)]
        return float(np.sum(vars_) / (len(delta_subset) ** 2)) if vars_ else np.nan

    rows = []
    protocols = list(scheme.enumerate_protocols())
    for k, (delta_sub, bsub) in enumerate(protocols):
        if verbose and k % 50 == 0:
            print(f"  protocol {k+1}/{len(protocols)}", flush=True)
        n_vol = scheme.volume_count(scheme.select_indices(delta_sub, bsub, include_b0=True))
        rec = {
            "n_deltas": len(delta_sub),
            "deltas": "+".join(str(int(d)) for d in delta_sub),
            "n_bvalues": len(bsub),
            "bvalues": "+".join(str(int(b)) for b in bsub),
            "n_volumes": n_vol,
            "scan_fraction": n_vol / scheme.full_volume_count(),
        }
        for which in ("MD", "FA"):
            # group comparison
            vh = [subject_roimean(s, delta_sub, bsub, which) for s in h_ids]
            vp = [subject_roimean(s, delta_sub, bsub, which) for s in p_ids]
            res = compare_groups(vh, vp, stats)
            rec[f"{which}_mean_healthy"] = res["mean_h"]
            rec[f"{which}_mean_patient"] = res["mean_p"]
            rec[f"{which}_U"] = res["U"]
            rec[f"{which}_p"] = res["p"]
            rec[f"{which}_cohens_d"] = res["cohens_d"]
            rec[f"{which}_rank_biserial"] = res["rank_biserial"]
            rec[f"{which}_sig05"] = bool(np.isfinite(res["p"]) and res["p"] < 0.05)

            # precision (CRLB SD per group)
            cvh = [v for v in (subject_crlb_var(s, delta_sub, bsub, which) for s in h_ids) if np.isfinite(v)]
            cvp = [v for v in (subject_crlb_var(s, delta_sub, bsub, which) for s in p_ids) if np.isfinite(v)]
            rec[f"{which}_CRLB_std_healthy"] = float(np.sqrt(np.mean(cvh))) if cvh else np.nan
            rec[f"{which}_CRLB_std_patient"] = float(np.sqrt(np.mean(cvp))) if cvp else np.nan

            # accuracy (RMSE vs full acquisition, over all subjects and per group)
            rall, rh, rp = [], [], []
            for s in subj:
                r = subject_rmse_vs_full(s, delta_sub, bsub, which)
                if np.isfinite(r):
                    rall.append(r)
                    (rh if group[s] == healthy_label else rp).append(r)
            rec[f"{which}_RMSE_vs_full"] = float(np.mean(rall)) if rall else np.nan
            rec[f"{which}_RMSE_vs_full_healthy"] = float(np.mean(rh)) if rh else np.nan
            rec[f"{which}_RMSE_vs_full_patient"] = float(np.mean(rp)) if rp else np.nan
        rows.append(rec)

    df = pd.DataFrame(rows)

    # reference verdict = the full protocol; flag protocols that agree with it
    full_mask = (df["n_deltas"] == len(all_deltas)) & (df["n_bvalues"] == len(scheme.nominal_bvalues))
    ref = df[full_mask].iloc[0]
    for which in ("MD", "FA"):
        df[f"{which}_matches_full"] = df[f"{which}_sig05"] == bool(ref[f"{which}_sig05"])

    # write outputs
    df.to_csv(os.path.join(outdir, "protocol_results.csv"), index=False)
    subj_info = pd.DataFrame([dict(subject_id=s, group=group[s],
                                   n_vox=subj[s]["n_vox"], sigma=subj[s]["sigma"]) for s in subj])
    _write_excel(df, subj_info, scheme, ref, os.path.join(outdir, "shortscan_results.xlsx"))
    _write_summary(df, scheme, ref, outdir)
    _make_heatmaps(df, scheme, outdir)
    if verbose:
        print(f"\nWrote {os.path.join(outdir, 'shortscan_results.xlsx')}")
    return df


# ---------------------------------------------------------------------- #
# Excel export (Notes / Subjects / ProtocolResults sheets)
# ---------------------------------------------------------------------- #
def _write_excel(df, subj_info, scheme, ref, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    base_font = Font(name=FONT, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_fmt = {
        "scan_fraction": "0.0%",
        "MD_mean_healthy": "0.000E+00", "MD_mean_patient": "0.000E+00",
        "MD_U": "0", "MD_p": "0.0000", "MD_cohens_d": "0.00", "MD_rank_biserial": "0.00",
        "MD_CRLB_std_healthy": "0.000E+00", "MD_CRLB_std_patient": "0.000E+00",
        "MD_RMSE_vs_full": "0.000E+00",
        "MD_RMSE_vs_full_healthy": "0.000E+00", "MD_RMSE_vs_full_patient": "0.000E+00",
        "FA_mean_healthy": "0.000", "FA_mean_patient": "0.000",
        "FA_U": "0", "FA_p": "0.0000", "FA_cohens_d": "0.00", "FA_rank_biserial": "0.00",
        "FA_CRLB_std_healthy": "0.0000", "FA_CRLB_std_patient": "0.0000",
        "FA_RMSE_vs_full": "0.0000",
        "FA_RMSE_vs_full_healthy": "0.0000", "FA_RMSE_vs_full_patient": "0.0000",
    }
    order = ["n_deltas", "deltas", "n_bvalues", "bvalues", "n_volumes", "scan_fraction",
             "MD_mean_healthy", "MD_mean_patient", "MD_U", "MD_p", "MD_cohens_d",
             "MD_rank_biserial", "MD_sig05", "MD_matches_full",
             "MD_CRLB_std_healthy", "MD_CRLB_std_patient",
             "MD_RMSE_vs_full", "MD_RMSE_vs_full_healthy", "MD_RMSE_vs_full_patient",
             "FA_mean_healthy", "FA_mean_patient", "FA_U", "FA_p", "FA_cohens_d",
             "FA_rank_biserial", "FA_sig05", "FA_matches_full",
             "FA_CRLB_std_healthy", "FA_CRLB_std_patient",
             "FA_RMSE_vs_full", "FA_RMSE_vs_full_healthy", "FA_RMSE_vs_full_patient"]
    order = [c for c in order if c in df.columns]
    dsheet = df[order].sort_values(["n_volumes", "deltas", "bvalues"])

    wb = Workbook()

    # Notes sheet: what the workbook contains + the statistical caveat
    ws = wb.active
    ws.title = "Notes"
    notes = [
        ("Shortened-scan DTI analysis (Study 2, report section 6)", True),
        ("", False),
        ("Full protocol: %d diffusion times x %d b-values = %d volumes."
         % (len(scheme.diffusion_times), len(scheme.nominal_bvalues), scheme.full_volume_count()), False),
        ("Each 'ProtocolResults' row is one shortened acquisition "
         "(subset of diffusion times x subset of b-values); %d rows." % len(df), False),
        ("", False),
        ("Columns:", True),
        ("  deltas / bvalues        diffusion times (ms) and non-zero b-values kept", False),
        ("  n_volumes / scan_fraction  volumes and fraction of the full 635-volume scan", False),
        ("  *_mean_healthy/patient   group-mean MD or FA (pooled over retained diffusion times)", False),
        ("  *_U, *_p                 two-sided Mann-Whitney U test between groups", False),
        ("  *_cohens_d, *_rank_biserial  effect sizes (healthy - patient)", False),
        ("  *_sig05                  TRUE if p < 0.05", False),
        ("  *_matches_full           TRUE if this protocol's p<0.05 verdict equals the full scan's", False),
        ("  *_CRLB_std_*             CRLB SD of MD/FA for this acquisition (smaller = more precise)", False),
        ("  *_RMSE_vs_full           voxelwise RMSE of MD/FA vs the full acquisition", False),
        ("", False),
        ("Units: MD and its CRLB/RMSE in mm^2/s; FA dimensionless.", False),
        ("", False),
        ("Caveat: with 5 healthy vs 4 patients the smallest two-sided Mann-Whitney p is 0.0159,", False),
        ("so p<0.05 is near-binary. Lean on effect sizes and *_matches_full, not p alone.", False),
        ("Full-protocol reference:  MD p=%.4g (d=%.2f);  FA p=%.4g (d=%.2f)."
         % (ref["MD_p"], ref["MD_cohens_d"], ref["FA_p"], ref["FA_cohens_d"]), False),
    ]
    for r, (text, bold) in enumerate(notes, start=1):
        ws.cell(row=r, column=1, value=text).font = Font(name=FONT, size=11, bold=bold)
    ws.column_dimensions["A"].width = 110

    # Subjects sheet
    ws2 = wb.create_sheet("Subjects")
    hdrs = list(subj_info.columns)
    for j, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal="center")
    for i, (_, r) in enumerate(subj_info.iterrows(), start=2):
        for j, h in enumerate(hdrs, 1):
            c = ws2.cell(row=i, column=j, value=r[h])
            c.font = base_font; c.border = border
            if h == "sigma":
                c.number_format = "0.000"
    for j, h in enumerate(hdrs, 1):
        ws2.column_dimensions[get_column_letter(j)].width = max(12, len(h) + 2)
    ws2.freeze_panes = "A2"

    # ProtocolResults sheet
    ws3 = wb.create_sheet("ProtocolResults")
    cols = list(dsheet.columns)
    for j, h in enumerate(cols, 1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, (_, r) in enumerate(dsheet.iterrows(), start=2):
        for j, h in enumerate(cols, 1):
            val = r[h]
            if isinstance(val, (np.bool_, bool)):
                val = bool(val)
            elif isinstance(val, np.integer):
                val = int(val)
            elif isinstance(val, np.floating):
                val = float(val)
            c = ws3.cell(row=i, column=j, value=val)
            c.font = base_font
            if h in col_fmt:
                c.number_format = col_fmt[h]
    for j, h in enumerate(cols, 1):
        ws3.column_dimensions[get_column_letter(j)].width = max(10, min(16, len(h) + 1))
    ws3.freeze_panes = "G2"
    ws3.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(dsheet) + 1)

    wb.save(path)


# ---------------------------------------------------------------------- #
def _write_summary(df, scheme, ref, outdir):
    lines = ["SHORTENED-SCAN ANALYSIS SUMMARY", "=" * 60,
             "Full protocol: %d diffusion times x %d b-values = %d volumes"
             % (len(scheme.diffusion_times), len(scheme.nominal_bvalues), scheme.full_volume_count())]
    for which in ("MD", "FA"):
        lines += ["", f"--- {which} ---",
                  "Full-protocol: p=%.4g, d=%.3f" % (ref[f"{which}_p"], ref[f"{which}_cohens_d"])]
        sig = df[df[f"{which}_sig05"]].sort_values(["n_volumes", f"{which}_p"])
        lines.append("Protocols with p<0.05: %d / %d" % (len(sig), len(df)))
        for _, r in sig.head(8).iterrows():
            lines.append("   d[%s] b[%s]  %d vol (%.0f%%)  p=%.4g d=%.2f  CRLB_std(H)=%.3g RMSE=%.3g"
                         % (r["deltas"], r["bvalues"], int(r["n_volumes"]), 100 * r["scan_fraction"],
                            r[f"{which}_p"], r[f"{which}_cohens_d"],
                            r[f"{which}_CRLB_std_healthy"], r[f"{which}_RMSE_vs_full"]))
    txt = "\n".join(lines)
    with open(os.path.join(outdir, "summary.txt"), "w") as f:
        f.write(txt + "\n")
    print("\n" + txt)


def _make_heatmaps(df, scheme, outdir):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    for which in ("MD", "FA"):
        piv = df.pivot_table(index="n_deltas", columns="n_bvalues",
                             values=f"{which}_sig05", aggfunc="mean")
        fig, ax = plt.subplots(figsize=(5, 4))
        im = ax.imshow(piv.values, origin="lower", aspect="auto", cmap="viridis", vmin=0, vmax=1)
        ax.set_xticks(range(len(piv.columns))); ax.set_xticklabels(piv.columns)
        ax.set_yticks(range(len(piv.index))); ax.set_yticklabels(piv.index)
        ax.set_xlabel("# b-values retained"); ax.set_ylabel("# diffusion times retained")
        ax.set_title(f"{which}: fraction of protocols with p<0.05")
        for i in range(piv.shape[0]):
            for j in range(piv.shape[1]):
                v = piv.values[i, j]
                if np.isfinite(v):
                    ax.text(j, i, f"{v:.2f}", ha="center", va="center",
                            color="white" if v < 0.6 else "black", fontsize=9)
        fig.colorbar(im, ax=ax, label="fraction p<0.05")
        fig.tight_layout()
        fig.savefig(os.path.join(outdir, f"heatmap_{which}.png"), dpi=130)
        plt.close(fig)


# ---------------------------------------------------------------------- #
# Synthetic self-test (no scans needed)
# ---------------------------------------------------------------------- #
def selftest(outdir):
    nib, gradient_table, TensorModel, stats = _import_heavy()
    os.makedirs(outdir, exist_ok=True)
    scheme = load_scheme("dwi.bval", "dwi.bvec")
    rng = np.random.default_rng(0)
    shape = (4, 4, 2)
    tmp = os.path.join(outdir, "synthetic"); os.makedirs(tmp, exist_ok=True)
    rows = []

    def make(sid, group):
        md0, fa0 = ((1.5e-3, 0.30) if group == "healthy" else (1.9e-3, 0.18))
        md0 += rng.normal(0, 0.06e-3); fa0 += rng.normal(0, 0.03)
        dwi = _simulate(scheme, shape, md0, fa0, rng, snr=40)
        dp = os.path.join(tmp, f"{sid}_dwi.nii.gz")
        mp = os.path.join(tmp, f"{sid}_mask.nii.gz")
        nib.save(nib.Nifti1Image(dwi.astype(np.float32), np.eye(4)), dp)
        nib.save(nib.Nifti1Image(np.ones(shape, np.int16), np.eye(4)), mp)
        rows.append(dict(subject_id=sid, group=group, dwi=dp, mask=mp))

    for i in range(5): make(f"C00{i+1}", "healthy")
    for i in [1, 2, 4, 5]: make(f"P00{i}", "patient")
    manifest = pd.DataFrame(rows)
    manifest.to_csv(os.path.join(outdir, "manifest.csv"), index=False)
    print("Synthetic data ready. Running analysis...\n")
    return run_analysis(manifest, scheme, "healthy", "patient", outdir)


def _simulate(scheme, shape, md, fa, rng, snr=40):
    from scipy.optimize import brentq
    def fa_of_k(k):
        l2 = 3 * md / (k + 2); l1 = k * l2
        lam = np.array([l1, l2, l2]); m = lam.mean()
        return np.sqrt(1.5 * ((lam - m) ** 2).sum() / (lam ** 2).sum())
    try:
        k = brentq(lambda k: fa_of_k(k) - fa, 1.0001, 60)
    except ValueError:
        k = 3.0
    l2 = 3 * md / (k + 2); l1 = k * l2
    D = np.diag([l1, l2, l2])
    g = scheme.bvecs.T; b = scheme.bvals; S0 = 1000.0
    sig = S0 * np.exp(-b * np.einsum("ij,jk,ik->i", g, D, g))
    vol = np.broadcast_to(sig, shape + (scheme.N,)).astype(float).copy()
    s = S0 / snr
    return np.sqrt((vol + rng.normal(0, s, vol.shape)) ** 2 + rng.normal(0, s, vol.shape) ** 2)


# ---------------------------------------------------------------------- #
def build_argparser():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--manifest", help="CSV: subject_id,group,dwi,mask")
    ap.add_argument("--data-root", help="root folder holding per-participant folders")
    ap.add_argument("--healthy-folders", default=",".join(DEFAULT_HEALTHY_FOLDERS))
    ap.add_argument("--patient-folders", default=",".join(DEFAULT_PATIENT_FOLDERS))
    ap.add_argument("--dwi-glob", default="*dwi*.nii*")
    ap.add_argument("--mask-glob", default="*mask*.nii*")
    ap.add_argument("--bval"); ap.add_argument("--bvec")
    ap.add_argument("--diffusion-times", default="70,130,190,250,330")
    ap.add_argument("--healthy-label", default="healthy")
    ap.add_argument("--patient-label", default="patient")
    ap.add_argument("--fit-method", default="WLS", choices=["WLS", "LS", "NLLS"])
    ap.add_argument("--outdir", default="results")
    ap.add_argument("--selftest", action="store_true")
    return ap


def main(argv=None):
    a = build_argparser().parse_args(argv)
    if a.selftest:
        selftest(a.outdir); return
    if not (a.bval and a.bvec):
        raise SystemExit("--bval and --bvec are required")
    dts = [float(x) for x in a.diffusion_times.split(",")]
    scheme = load_scheme(a.bval, a.bvec, diffusion_times=dts)
    if a.manifest:
        manifest = pd.read_csv(a.manifest)
    elif a.data_root:
        manifest = build_manifest_from_root(
            a.data_root, a.healthy_folders.split(","), a.patient_folders.split(","),
            a.healthy_label, a.patient_label, dwi_glob=a.dwi_glob, mask_glob=a.mask_glob)
        print("Manifest:\n" + manifest.to_string(index=False))
    else:
        raise SystemExit("provide --manifest or --data-root")
    run_analysis(manifest, scheme, a.healthy_label, a.patient_label, a.outdir, fit_method=a.fit_method)


if __name__ == "__main__":
    main()
